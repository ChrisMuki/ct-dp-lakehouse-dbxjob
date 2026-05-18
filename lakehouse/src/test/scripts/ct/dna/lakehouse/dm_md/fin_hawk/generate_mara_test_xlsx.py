#!/usr/bin/env python3
"""Regenerate MaraTest.xlsx for ct.dna.lakehouse.dm_md.fin_hawk.MaraTest.

Run via:  python3 lakehouse/src/test/scripts/ct/dna/lakehouse/dm_md/fin_hawk/generate_mara_test_xlsx.py

------------------------------------------------------------------------------
Test design — per-row transition matrix for the simple-upsert dm_md.mara merge.
------------------------------------------------------------------------------

Unlike makt, mara has no per-language pivot: it just CDF-merges each per-PK
slice. We enumerate 7 transitions a single source row can undergo between two
snapshots ("pre" → "post"):

   1  value v1     → value v2                     (CDF update)
   2  value v1     → NULL                         (CDF update, value cleared)
   3  value v1     → same v1 but alt_col differs  (CDF update, mtart unchanged)
   4  value v1     → row bit-identical            (no CDF event)
   5  value v1     → row deleted                  (CDF delete)
   6  no row       → value                        (CDF insert)
   7  no row       → value NULL                   (CDF insert with NULL mtart)

Phases mirror MaktTest:

  • Phase 0 — every source and the target are empty. update(mara) must leave
    target empty.
  • Phase 1 — e32 and epp both receive their "pre" snapshots; every row is a
    CDF insert. The other 12 sources stay empty (feeds report isUnchanged
    and contribute nothing).
  • Phase 2 — e32 is MERGE-aligned to "post"; pre→post diff drives CDF
    insert/update/delete per row. EPP must stay bit-identical.
  • Phase 3 — EPP is dropped, recreated and re-loaded with "post". Because
    its previously-recorded known commit no longer exists, the framework
    treats the feed as a snapshot. Snapshot semantics: rows absent from
    snapshot are deleted via whenNotMatchedBySource (gated on
    _mk_system ∈ snapshotSystems); E32 rows must stay untouched.

Source rows are ~270 columns wide each; the test only varies `mtart` (a
column the merge consumes), and `_mk_created_at` for the alt_col case to
trigger a CDF event without changing `mtart`. All other columns are NULL
across pre and post — DecimalType / BinaryType cells aren't supported by
TestDataManager.read so they MUST stay NULL.
"""
from datetime import datetime
from pathlib import Path

from openpyxl import LXML
from openpyxl import Workbook
from openpyxl.cell import _writer as _cell_writer
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet import _writer as _ws_writer
from openpyxl.xml.functions import Element, SubElement, XML_NS

EMPTY_STRING_FILL = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")


def patch_openpyxl_empty_string_cells() -> None:
    """Monkey-patch openpyxl so empty-string cells are written as real
    `<is><t></t></is>` inline strings (POI reads them as ""), not bare `<c/>`
    (POI reads them as BLANK → Spark NULL). See TEST_XLSX_AUTHORING.md §2.
    """
    if getattr(_cell_writer, "_EMPTY_STRING_PATCH_APPLIED", False):
        return

    def _patched_etree_write_cell(xf, worksheet, cell, styled=None):
        value, attributes = _cell_writer._set_attributes(cell, styled)
        el = Element("c", attributes)
        if cell.data_type == "s":
            is_el = SubElement(el, "is")
            t_el = SubElement(is_el, "t")
            if value:
                t_el.text = value
            if value is not None and value != value.strip():
                t_el.set("{%s}space" % XML_NS, "preserve")
            xf.write(el)
            return
        _cell_writer._ORIGINAL_etree_write_cell(xf, worksheet, cell, styled)

    def _patched_lxml_write_cell(xf, worksheet, cell, styled=False):
        value, attributes = _cell_writer._set_attributes(cell, styled)
        if cell.data_type == "s":
            with xf.element("c", attributes):
                with xf.element("is"):
                    attrs = {}
                    if value is not None and value != value.strip():
                        attrs["{%s}space" % XML_NS] = "preserve"
                    el = Element("t", attrs)
                    el.text = value or ""
                    xf.write(el)
            return
        _cell_writer._ORIGINAL_lxml_write_cell(xf, worksheet, cell, styled)

    _cell_writer._ORIGINAL_etree_write_cell = _cell_writer.etree_write_cell
    _cell_writer._ORIGINAL_lxml_write_cell = _cell_writer.lxml_write_cell
    _cell_writer.etree_write_cell = _patched_etree_write_cell
    _cell_writer.lxml_write_cell = _patched_lxml_write_cell
    _cell_writer.write_cell = _patched_lxml_write_cell if LXML else _patched_etree_write_cell
    _ws_writer.write_cell = _cell_writer.write_cell
    _cell_writer._EMPTY_STRING_PATCH_APPLIED = True


patch_openpyxl_empty_string_cells()


def _resource_sibling(filename: str) -> Path:
    here = Path(__file__).resolve()
    scripts_root = next(p for p in here.parents if p.name == "scripts" and p.parent.name == "test")
    pkg = here.parent.relative_to(scripts_root)
    return scripts_root.parent / "resources" / pkg / filename


OUT = _resource_sibling("MaraTest.xlsx")


# ---------------------------------------------------------------------------
# Embedded source schemas (ordered as ct_gbl_*.mara.entityStructType).
# Extracted from `lakehouse/testOnly ct.dna.lakehouse.dm_md.fin_hawk.PrintMaraSchemas`.
# Format: (name, dtype). dtype matches Spark `DataType.simpleString` output.
#
# Important: TestDataManager.read only supports cell types
#   string / int / long / double / boolean / date / timestamp / struct
# Decimal, Binary, etc. columns must remain NULL in every row, or read will
# throw "Unsupported CellType".
E32_FIELDS = [
    ("mandt", "string"),
    ("matnr", "string"),
    ("ersda", "string"),
    ("ernam", "string"),
    ("laeda", "string"),
    ("aenam", "string"),
    ("vpsta", "string"),
    ("pstat", "string"),
    ("lvorm", "string"),
    ("mtart", "string"),
    ("mbrsh", "string"),
    ("matkl", "string"),
    ("bismt", "string"),
    ("meins", "string"),
    ("bstme", "string"),
    ("zeinr", "string"),
    ("zeiar", "string"),
    ("zeivr", "string"),
    ("zeifo", "string"),
    ("aeszn", "string"),
    ("blatt", "string"),
    ("blanz", "string"),
    ("ferth", "string"),
    ("formt", "string"),
    ("groes", "string"),
    ("wrkst", "string"),
    ("normt", "string"),
    ("labor", "string"),
    ("ekwsl", "string"),
    ("brgew", "decimal(13,3)"),
    ("ntgew", "decimal(13,3)"),
    ("gewei", "string"),
    ("volum", "decimal(13,3)"),
    ("voleh", "string"),
    ("behvo", "string"),
    ("raube", "string"),
    ("tempb", "string"),
    ("disst", "string"),
    ("tragr", "string"),
    ("stoff", "string"),
    ("spart", "string"),
    ("kunnr", "string"),
    ("eannr", "string"),
    ("wesch", "decimal(13,3)"),
    ("bwvor", "string"),
    ("bwscl", "string"),
    ("saiso", "string"),
    ("etiar", "string"),
    ("etifo", "string"),
    ("entar", "string"),
    ("ean11", "string"),
    ("numtp", "string"),
    ("laeng", "decimal(13,3)"),
    ("breit", "decimal(13,3)"),
    ("hoehe", "decimal(13,3)"),
    ("meabm", "string"),
    ("prdha", "string"),
    ("aeklk", "string"),
    ("cadkz", "string"),
    ("qmpur", "string"),
    ("ergew", "decimal(13,3)"),
    ("ergei", "string"),
    ("ervol", "decimal(13,3)"),
    ("ervoe", "string"),
    ("gewto", "decimal(3,1)"),
    ("volto", "decimal(3,1)"),
    ("vabme", "string"),
    ("kzrev", "string"),
    ("kzkfg", "string"),
    ("xchpf", "string"),
    ("vhart", "string"),
    ("fuelg", "decimal(3,0)"),
    ("stfak", "int"),
    ("magrv", "string"),
    ("begru", "string"),
    ("datab", "string"),
    ("liqdt", "string"),
    ("saisj", "string"),
    ("plgtp", "string"),
    ("mlgut", "string"),
    ("extwg", "string"),
    ("satnr", "string"),
    ("attyp", "string"),
    ("kzkup", "string"),
    ("kznfm", "string"),
    ("pmata", "string"),
    ("mstae", "string"),
    ("mstav", "string"),
    ("mstde", "string"),
    ("mstdv", "string"),
    ("taklv", "string"),
    ("rbnrm", "string"),
    ("mhdrz", "decimal(4,0)"),
    ("mhdhb", "decimal(4,0)"),
    ("mhdlp", "decimal(3,0)"),
    ("inhme", "string"),
    ("inhal", "decimal(13,3)"),
    ("vpreh", "decimal(5,0)"),
    ("etiag", "string"),
    ("inhbr", "decimal(13,3)"),
    ("cmeth", "string"),
    ("cuobf", "string"),
    ("kzumw", "string"),
    ("kosch", "string"),
    ("sprof", "string"),
    ("nrfhg", "string"),
    ("mfrpn", "string"),
    ("mfrnr", "string"),
    ("bmatn", "string"),
    ("mprof", "string"),
    ("kzwsm", "string"),
    ("saity", "string"),
    ("profl", "string"),
    ("ihivi", "string"),
    ("iloos", "string"),
    ("serlv", "string"),
    ("kzgvh", "string"),
    ("xgchp", "string"),
    ("kzeff", "string"),
    ("compl", "string"),
    ("iprkz", "string"),
    ("rdmhd", "string"),
    ("przus", "string"),
    ("mtpos_mara", "string"),
    ("bflme", "string"),
    ("matfi", "string"),
    ("cmrel", "string"),
    ("bbtyp", "string"),
    ("sled_bbd", "string"),
    ("gtin_variant", "string"),
    ("gennr", "string"),
    ("rmatp", "string"),
    ("gds_relevant", "string"),
    ("weora", "string"),
    ("hutyp_dflt", "string"),
    ("pilferable", "string"),
    ("whstc", "string"),
    ("whmatgr", "string"),
    ("hndlcode", "string"),
    ("hazmat", "string"),
    ("hutyp", "string"),
    ("tare_var", "string"),
    ("maxc", "decimal(15,3)"),
    ("maxc_tol", "decimal(3,1)"),
    ("maxl", "decimal(15,3)"),
    ("maxb", "decimal(15,3)"),
    ("maxh", "decimal(15,3)"),
    ("maxdim_uom", "string"),
    ("herkl", "string"),
    ("mfrgr", "string"),
    ("qqtime", "decimal(3,0)"),
    ("qqtimeuom", "string"),
    ("qgrp", "string"),
    ("serial", "string"),
    ("ps_smartform", "string"),
    ("logunit", "string"),
    ("cwqrel", "string"),
    ("cwqproc", "string"),
    ("cwqtolgr", "string"),
    ("adprof", "string"),
    ("ipmipproduct", "string"),
    ("allow_pmat_igno", "string"),
    ("medium", "string"),
    ("commodity", "string"),
    ("animal_origin", "string"),
    ("textile_comp_ind", "string"),
    ("sgt_csgr", "string"),
    ("sgt_covsa", "string"),
    ("sgt_stat", "string"),
    ("sgt_scope", "string"),
    ("sgt_rel", "string"),
    ("anp", "string"),
    ("fsh_mg_at1", "string"),
    ("fsh_mg_at2", "string"),
    ("fsh_mg_at3", "string"),
    ("fsh_sealv", "string"),
    ("fsh_seaim", "string"),
    ("fsh_sc_mid", "string"),
    ("psm_code", "string"),
    ("_bev1_luleinh", "string"),
    ("_bev1_luldegrp", "string"),
    ("_bev1_nestruccat", "string"),
    ("_dsd_sl_toltyp", "string"),
    ("_dsd_sv_cnt_grp", "string"),
    ("_dsd_vc_group", "string"),
    ("_vso_r_tilt_ind", "string"),
    ("_vso_r_stack_ind", "string"),
    ("_vso_r_bot_ind", "string"),
    ("_vso_r_top_ind", "string"),
    ("_vso_r_stack_no", "string"),
    ("_vso_r_pal_ind", "string"),
    ("_vso_r_pal_ovr_d", "decimal(13,3)"),
    ("_vso_r_pal_ovr_w", "decimal(13,3)"),
    ("_vso_r_pal_b_ht", "decimal(13,3)"),
    ("_vso_r_pal_min_h", "decimal(13,3)"),
    ("_vso_r_tol_b_ht", "decimal(13,3)"),
    ("_vso_r_no_p_gvh", "string"),
    ("_vso_r_quan_unit", "string"),
    ("_vso_r_kzgvh_ind", "string"),
    ("packcode", "string"),
    ("dg_pack_status", "string"),
    ("mcond", "string"),
    ("retdelc", "string"),
    ("loglev_reto", "string"),
    ("nsnid", "string"),
    ("adspc_spc", "string"),
    ("imatn", "string"),
    ("picnum", "string"),
    ("bstat", "string"),
    ("color_atinn", "string"),
    ("size1_atinn", "string"),
    ("size2_atinn", "string"),
    ("color", "string"),
    ("size1", "string"),
    ("size2", "string"),
    ("free_char", "string"),
    ("care_code", "string"),
    ("brand_id", "string"),
    ("fiber_code1", "string"),
    ("fiber_part1", "string"),
    ("fiber_code2", "string"),
    ("fiber_part2", "string"),
    ("fiber_code3", "string"),
    ("fiber_part3", "string"),
    ("fiber_code4", "string"),
    ("fiber_part4", "string"),
    ("fiber_code5", "string"),
    ("fiber_part5", "string"),
    ("fashgrd", "string"),
    ("zz_matnr_neu", "string"),
    ("zz_kdmat_neu", "string"),
    ("zz_kdmat_txt_neu", "string"),
    ("zz_zusatztext", "string"),
    ("zz_praegung", "string"),
    ("zz_dicke", "string"),
    ("zz_anz_etik", "decimal(3,0)"),
    ("zz_abrdatum", "string"),
    ("zz_konfekt", "string"),
    ("zz_ernam_etik", "string"),
    ("zz_erdat_etik", "string"),
    ("zz_erzeit_etik", "string"),
    ("zz_dessin", "string"),
    ("_mk_org", "string"),
    ("_mk_site", "string"),
    ("_mk_system", "string"),
    ("_mk_instance", "string"),
    ("_mk_partition", "string"),
    ("_mk_file", "string"),
    ("_mk_container", "string"),
    ("_mk_account", "string"),
    ("_mk_created_at", "timestamp"),
    ("_lh_id_in_message", "bigint"),
    ("_lh_ingest_warning", "string"),
    ("zz_farbe_string", "string"),
    ("zz_laenge_string", "string"),
    ("zz_breite_string", "string"),
    ("zz_format_string", "string"),
    ("zz_praefix_matnr_string", "string"),
    ("zz_hd_vererben_string", "string"),
    ("zz_desfam_string", "string"),
    ("zz_konform_string", "string"),
    ("zz_dat_pr_string", "string"),
    ("zz_dat_pvo_string", "string"),
    ("zzgold_matnr_string", "string"),
    ("zzgold_maktx_string", "string"),
    ("zzgold_sysid_string", "string"),
    ("zzgold_mandt_string", "string"),
]
EPP_FIELDS = [
    ("mandt", "string"),
    ("matnr", "string"),
    ("ersda", "string"),
    ("ernam", "string"),
    ("laeda", "string"),
    ("aenam", "string"),
    ("vpsta", "string"),
    ("pstat", "string"),
    ("lvorm", "string"),
    ("mtart", "string"),
    ("mbrsh", "string"),
    ("matkl", "string"),
    ("bismt", "string"),
    ("meins", "string"),
    ("bstme", "string"),
    ("zeinr", "string"),
    ("zeiar", "string"),
    ("zeivr", "string"),
    ("zeifo", "string"),
    ("aeszn", "string"),
    ("blatt", "string"),
    ("blanz", "string"),
    ("ferth", "string"),
    ("formt", "string"),
    ("groes", "string"),
    ("wrkst", "string"),
    ("normt", "string"),
    ("labor", "string"),
    ("ekwsl", "string"),
    ("brgew", "decimal(13,3)"),
    ("ntgew", "decimal(13,3)"),
    ("gewei", "string"),
    ("volum", "decimal(13,3)"),
    ("voleh", "string"),
    ("behvo", "string"),
    ("raube", "string"),
    ("tempb", "string"),
    ("disst", "string"),
    ("tragr", "string"),
    ("stoff", "string"),
    ("spart", "string"),
    ("kunnr", "string"),
    ("eannr", "string"),
    ("wesch", "decimal(13,3)"),
    ("bwvor", "string"),
    ("bwscl", "string"),
    ("saiso", "string"),
    ("etiar", "string"),
    ("etifo", "string"),
    ("entar", "string"),
    ("ean11", "string"),
    ("numtp", "string"),
    ("laeng", "decimal(13,3)"),
    ("breit", "decimal(13,3)"),
    ("hoehe", "decimal(13,3)"),
    ("meabm", "string"),
    ("prdha", "string"),
    ("aeklk", "string"),
    ("cadkz", "string"),
    ("qmpur", "string"),
    ("ergew", "decimal(13,3)"),
    ("ergei", "string"),
    ("ervol", "decimal(13,3)"),
    ("ervoe", "string"),
    ("gewto", "decimal(3,1)"),
    ("volto", "decimal(3,1)"),
    ("vabme", "string"),
    ("kzrev", "string"),
    ("kzkfg", "string"),
    ("xchpf", "string"),
    ("vhart", "string"),
    ("fuelg", "decimal(3,0)"),
    ("stfak", "int"),
    ("magrv", "string"),
    ("begru", "string"),
    ("datab", "string"),
    ("liqdt", "string"),
    ("saisj", "string"),
    ("plgtp", "string"),
    ("mlgut", "string"),
    ("extwg", "string"),
    ("satnr", "string"),
    ("attyp", "string"),
    ("kzkup", "string"),
    ("kznfm", "string"),
    ("pmata", "string"),
    ("mstae", "string"),
    ("mstav", "string"),
    ("mstde", "string"),
    ("mstdv", "string"),
    ("taklv", "string"),
    ("rbnrm", "string"),
    ("mhdrz", "decimal(4,0)"),
    ("mhdhb", "decimal(4,0)"),
    ("mhdlp", "decimal(3,0)"),
    ("inhme", "string"),
    ("inhal", "decimal(13,3)"),
    ("vpreh", "decimal(5,0)"),
    ("etiag", "string"),
    ("inhbr", "decimal(13,3)"),
    ("cmeth", "string"),
    ("cuobf", "string"),
    ("kzumw", "string"),
    ("kosch", "string"),
    ("sprof", "string"),
    ("nrfhg", "string"),
    ("mfrpn", "string"),
    ("mfrnr", "string"),
    ("bmatn", "string"),
    ("mprof", "string"),
    ("kzwsm", "string"),
    ("saity", "string"),
    ("profl", "string"),
    ("ihivi", "string"),
    ("iloos", "string"),
    ("serlv", "string"),
    ("kzgvh", "string"),
    ("xgchp", "string"),
    ("kzeff", "string"),
    ("compl", "string"),
    ("iprkz", "string"),
    ("rdmhd", "string"),
    ("przus", "string"),
    ("mtpos_mara", "string"),
    ("bflme", "string"),
    ("matfi", "string"),
    ("cmrel", "string"),
    ("bbtyp", "string"),
    ("sled_bbd", "string"),
    ("gtin_variant", "string"),
    ("gennr", "string"),
    ("rmatp", "string"),
    ("gds_relevant", "string"),
    ("weora", "string"),
    ("hutyp_dflt", "string"),
    ("pilferable", "string"),
    ("whstc", "string"),
    ("whmatgr", "string"),
    ("hndlcode", "string"),
    ("hazmat", "string"),
    ("hutyp", "string"),
    ("tare_var", "string"),
    ("maxc", "decimal(15,3)"),
    ("maxc_tol", "decimal(3,1)"),
    ("maxl", "decimal(15,3)"),
    ("maxb", "decimal(15,3)"),
    ("maxh", "decimal(15,3)"),
    ("maxdim_uom", "string"),
    ("herkl", "string"),
    ("mfrgr", "string"),
    ("qqtime", "decimal(3,0)"),
    ("qqtimeuom", "string"),
    ("qgrp", "string"),
    ("serial", "string"),
    ("ps_smartform", "string"),
    ("logunit", "string"),
    ("cwqrel", "string"),
    ("cwqproc", "string"),
    ("cwqtolgr", "string"),
    ("adprof", "string"),
    ("ipmipproduct", "string"),
    ("allow_pmat_igno", "string"),
    ("medium", "string"),
    ("commodity", "string"),
    ("animal_origin", "string"),
    ("textile_comp_ind", "string"),
    ("sgt_csgr", "string"),
    ("sgt_covsa", "string"),
    ("sgt_stat", "string"),
    ("sgt_scope", "string"),
    ("sgt_rel", "string"),
    ("anp", "string"),
    ("fsh_mg_at1", "string"),
    ("fsh_mg_at2", "string"),
    ("fsh_mg_at3", "string"),
    ("fsh_sealv", "string"),
    ("fsh_seaim", "string"),
    ("fsh_sc_mid", "string"),
    ("psm_code", "string"),
    ("_bev1_luleinh", "string"),
    ("_bev1_luldegrp", "string"),
    ("_bev1_nestruccat", "string"),
    ("_dsd_sl_toltyp", "string"),
    ("_dsd_sv_cnt_grp", "string"),
    ("_dsd_vc_group", "string"),
    ("_sapmp_kadu", "decimal(7,2)"),
    ("_sapmp_abmein", "string"),
    ("_sapmp_kadp", "decimal(5,2)"),
    ("_sapmp_brad", "decimal(4,0)"),
    ("_sapmp_spbi", "decimal(7,2)"),
    ("_sapmp_trad", "decimal(7,2)"),
    ("_sapmp_kedu", "decimal(7,2)"),
    ("_sapmp_sptr", "decimal(13,3)"),
    ("_sapmp_fbdk", "decimal(7,2)"),
    ("_sapmp_fbhk", "decimal(7,2)"),
    ("_sapmp_rili", "string"),
    ("_sapmp_fbak", "string"),
    ("_sapmp_aho", "string"),
    ("_sapmp_mifrr", "decimal(7,2)"),
    ("_vso_r_tilt_ind", "string"),
    ("_vso_r_stack_ind", "string"),
    ("_vso_r_bot_ind", "string"),
    ("_vso_r_top_ind", "string"),
    ("_vso_r_stack_no", "string"),
    ("_vso_r_pal_ind", "string"),
    ("_vso_r_pal_ovr_d", "decimal(13,3)"),
    ("_vso_r_pal_ovr_w", "decimal(13,3)"),
    ("_vso_r_pal_b_ht", "decimal(13,3)"),
    ("_vso_r_pal_min_h", "decimal(13,3)"),
    ("_vso_r_tol_b_ht", "decimal(13,3)"),
    ("_vso_r_no_p_gvh", "string"),
    ("_vso_r_quan_unit", "string"),
    ("_vso_r_kzgvh_ind", "string"),
    ("packcode", "string"),
    ("dg_pack_status", "string"),
    ("mcond", "string"),
    ("retdelc", "string"),
    ("loglev_reto", "string"),
    ("nsnid", "string"),
    ("ovlpn", "string"),
    ("adspc_spc", "string"),
    ("msbookpartno", "string"),
    ("vtype", "string"),
    ("vchnr", "string"),
    ("evval", "string"),
    ("dvers", "string"),
    ("dpcbt", "string"),
    ("xgrdt", "string"),
    ("imatn", "string"),
    ("picnum", "string"),
    ("bstat", "string"),
    ("color_atinn", "string"),
    ("size1_atinn", "string"),
    ("size2_atinn", "string"),
    ("color", "string"),
    ("size1", "string"),
    ("size2", "string"),
    ("free_char", "string"),
    ("care_code", "string"),
    ("brand_id", "string"),
    ("fiber_code1", "string"),
    ("_mk_org", "string"),
    ("_mk_site", "string"),
    ("_mk_system", "string"),
    ("_mk_instance", "string"),
    ("_mk_partition", "string"),
    ("_mk_file", "string"),
    ("_mk_container", "string"),
    ("_mk_account", "string"),
    ("_mk_created_at", "timestamp"),
    ("_lh_id_in_message", "bigint"),
    ("_lh_ingest_warning", "string"),
    ("varid_binary", "binary"),
    ("fiber_part1_string", "string"),
    ("fiber_code2_string", "string"),
    ("fiber_part2_string", "string"),
    ("fiber_code3_string", "string"),
    ("fiber_part3_string", "string"),
    ("fiber_code4_string", "string"),
    ("fiber_part4_string", "string"),
    ("fiber_code5_string", "string"),
    ("fiber_part5_string", "string"),
    ("fashgrd_string", "string"),
    ("zzstcc_string", "string"),
    ("yytarif_string", "string"),
    ("yyaland_string", "string"),
    ("yyprocu_string", "string"),
    ("yynafta_string", "string"),
    ("yydtydrw_string", "string"),
    ("yycsiscde_string", "string"),
    ("ycomty_string", "string"),
    ("yprism_string", "string"),
    ("zzgold_matnr_string", "string"),
    ("zzgold_maktx_string", "string"),
    ("zzgold_sysid_string", "string"),
    ("zzgold_mandt_string", "string"),
]

E32_NAMES = [n for n, _ in E32_FIELDS]
EPP_NAMES = [n for n, _ in EPP_FIELDS]
E32_TYPES = dict(E32_FIELDS)
EPP_TYPES = dict(EPP_FIELDS)


# ---------------------------------------------------------------------------
# Target schema (dm_md.fin_hawk.mara.entityStructType).
# ---------------------------------------------------------------------------
TGT_HEADER = [
    "_mk_system", "_mk_instance", "matnr",
    "mtart", "matkl", "ersda", "pstat", "vpsta", "lvorm", "meins",
    "ferth", "formt", "groes", "wrkst", "normt",
    "brgew", "ntgew", "gewei", "volum", "voleh",
    "laeng", "breit", "hoehe", "meabm", "prdha", "attyp", "mfrpn", "mfrnr",
]


# ---------------------------------------------------------------------------
# Case matrix. We vary `mtart` (a consumed value column) per row.
# `alt_col=True` forces _mk_created_at to differ between pre and post so the
# CDF emits an UPDATE event even when mtart is unchanged.
# ---------------------------------------------------------------------------
CASE_DESCRIPTIONS = {
    1: "value -> other value",
    2: "value -> NULL",
    3: "value -> same value (alt col changed)",
    4: "row bit-identical (no event)",
    5: "value -> row deleted",
    6: "no row -> value",
    7: "no row -> value NULL",
}

CASES = {
    1: dict(pre_row=True,  pre_mtart="pre",  post_row=True,  post_mtart="post", alt_col=False),
    2: dict(pre_row=True,  pre_mtart="pre",  post_row=True,  post_mtart=None,   alt_col=False),
    3: dict(pre_row=True,  pre_mtart="same", post_row=True,  post_mtart="same", alt_col=True),
    4: dict(pre_row=True,  pre_mtart="same", post_row=True,  post_mtart="same", alt_col=False),
    5: dict(pre_row=True,  pre_mtart="pre",  post_row=False, post_mtart=None,   alt_col=False),
    6: dict(pre_row=False, pre_mtart=None,   post_row=True,  post_mtart="post", alt_col=False),
    7: dict(pre_row=False, pre_mtart=None,   post_row=True,  post_mtart=None,   alt_col=False),
}
ALL_CASES = sorted(CASES.keys())

TS_PRE  = datetime(2024, 1, 1, 12, 0, 0)
TS_POST = datetime(2024, 6, 1, 12, 0, 0)

INSTANCE = "100"
ORG  = "CT"
SITE = "GBL"

ABSENT = object()


def matnr_for(i): return f"M{i:02d}"


def mtart_value(case_id, snapshot):
    c = CASES[case_id]
    present = c["pre_row"] if snapshot == "pre" else c["post_row"]
    if not present:
        return ABSENT
    raw = c["pre_mtart"] if snapshot == "pre" else c["post_mtart"]
    if raw is None:
        return None
    return f"mtart_{matnr_for(case_id)}_{raw}"


def ts_value(case_id, snapshot):
    c = CASES[case_id]
    if not c["alt_col"]:
        return TS_PRE
    return TS_PRE if snapshot == "pre" else TS_POST


# ---------------------------------------------------------------------------
# Source row builder. Returns one row of full source-entity width with
# defaults applied; only consumed/identity columns are overridden.
# ---------------------------------------------------------------------------
def build_src_row(field_names, system, matnr, mtart, ts, case_desc):
    """Build a single source-entity row. All fields default to None; we then
    overwrite the metadata fields and `mtart` only. The leading `#case`
    comment column is appended once at the workbook-write level.
    """
    overrides = {
        "mandt":             "100",
        "matnr":             matnr,
        "mtart":             mtart,  # may be None
        "_mk_org":           ORG,
        "_mk_site":          SITE,
        "_mk_system":        system,
        "_mk_instance":      INSTANCE,
        "_mk_partition":     "p",
        "_mk_file":          "f",
        "_mk_container":     "c",
        "_mk_account":       "a",
        "_mk_created_at":    ts,
        "_lh_id_in_message": 1,
    }
    row = [overrides.get(name, None) for name in field_names]
    return [case_desc] + row


def build_source_rows(field_names, system, snapshot):
    """Generate rows for one (system, snapshot) of the case matrix."""
    rows = []
    for i in ALL_CASES:
        matnr = matnr_for(i)
        case_desc = f"c{i:02d}: {CASE_DESCRIPTIONS[i]}"
        mx = mtart_value(i, snapshot)
        if mx is ABSENT:
            continue
        rows.append(build_src_row(field_names, system, matnr, mx, ts_value(i, snapshot), case_desc))
    return rows


# ---------------------------------------------------------------------------
# Merge simulator (mirrors mara.executeTransaction one-to-one).
#
#   target state per system: { matnr -> mtart }   (other value cols stay None,
#   identical pre/post — see module docstring; not tracked here).
#
# pivot per matnr (from a CDF feed):
#   None              → no event, leave target alone
#   ("delete", None)  → whenMatched(isDelete).delete()
#   ("upsert", v)     → whenMatched().update(...) or whenNotMatched(isUpsert).insert(...)
#                        with mtart=v (v may be None)
#
# snapshot feed:
#   contains exactly the post-state row for every present matnr; absent matnrs
#   are deleted via whenNotMatchedBySource(target._mk_system isin snapshotSystems).
# ---------------------------------------------------------------------------
def cdf_event(case_id, snapshot_kind):
    """Return the per-row CDF event for one case under one feed mode.

    snapshot_kind ∈ {"initial", "delta"}:
      initial — source went empty→pre (every pre row is an insert).
      delta   — source went pre→post (per-row diff).
    """
    c = CASES[case_id]

    if snapshot_kind == "initial":
        if not c["pre_row"]:
            return None
        return ("upsert", None if c["pre_mtart"] is None
                else f"mtart_{matnr_for(case_id)}_{c['pre_mtart']}")

    # delta
    if not c["pre_row"] and not c["post_row"]:
        return None  # neither pre nor post → no event
    if c["pre_row"] and not c["post_row"]:
        return ("delete", None)
    if not c["pre_row"] and c["post_row"]:
        return ("upsert", None if c["post_mtart"] is None
                else f"mtart_{matnr_for(case_id)}_{c['post_mtart']}")
    same_mtart = c["pre_mtart"] == c["post_mtart"]
    if same_mtart and not c["alt_col"]:
        return None  # row bit-identical → no event
    return ("upsert", None if c["post_mtart"] is None
            else f"mtart_{matnr_for(case_id)}_{c['post_mtart']}")


def snapshot_event(case_id):
    """Snapshot feed: every post-present row appears with its post value;
    absent rows trigger whenNotMatchedBySource at the per-feed level."""
    c = CASES[case_id]
    if not c["post_row"]:
        return None
    return ("upsert", None if c["post_mtart"] is None
            else f"mtart_{matnr_for(case_id)}_{c['post_mtart']}")


def apply_phase(state_per_system, system, mode):
    prev = state_per_system.get(system, {})
    new_state = dict(prev)
    is_snapshot = mode == "snapshot"

    present_matnrs_for_snapshot = set()
    for i in ALL_CASES:
        matnr = matnr_for(i)
        if mode == "snapshot":
            ev = snapshot_event(i)
        else:
            ev = cdf_event(i, mode)
        if ev is None:
            continue
        kind, value = ev
        if kind == "delete":
            new_state.pop(matnr, None)
        else:  # upsert (insert or update)
            new_state[matnr] = value
            present_matnrs_for_snapshot.add(matnr)

    if is_snapshot:
        # whenNotMatchedBySource on this system's rows — delete any matnr
        # that was in the previous target but not in the snapshot.
        for matnr in list(new_state.keys()):
            if matnr not in present_matnrs_for_snapshot:
                # only this system's rows can be deleted, but state_per_system
                # is partitioned per-system so we're already in that scope.
                del new_state[matnr]

    state_per_system[system] = new_state


def state_to_rows(state_per_system):
    rows = []
    for system in sorted(state_per_system.keys()):
        for matnr in sorted(state_per_system[system].keys()):
            mtart = state_per_system[system][matnr]
            rows.append([
                system, INSTANCE, matnr,
                mtart,
                None, None, None, None, None, None,       # matkl..meins
                None, None, None, None, None,             # ferth..normt
                None, None, None, None, None,             # brgew..voleh
                None, None, None, None, None, None, None, None,  # laeng..mfrnr
            ])
    return rows


# ---------------------------------------------------------------------------
# Workbook assembly.
# ---------------------------------------------------------------------------
def add_sheet(wb, name, header, rows):
    ws = wb.create_sheet(name)
    ws.append(header)
    for r in rows:
        ws.append(r)
        excel_row_idx = ws.max_row
        for col_idx, value in enumerate(r, start=1):
            if value == "":
                cell = ws.cell(row=excel_row_idx, column=col_idx)
                cell.value = ""
                cell.data_type = "s"
                cell.quotePrefix = True

    if rows:
        last_col = get_column_letter(len(header))
        data_range = f"A2:{last_col}{ws.max_row}"
        rule = FormulaRule(
            formula=['AND(NOT(ISBLANK(A2)), LEN(A2)=0)'],
            fill=EMPTY_STRING_FILL,
        )
        ws.conditional_formatting.add(data_range, rule)


wb = Workbook()
del wb["Sheet"]

E32_HEADER = ["#case"] + E32_NAMES
EPP_HEADER = ["#case"] + EPP_NAMES

e32_pre  = build_source_rows(E32_NAMES, "E32", "pre")
e32_post = build_source_rows(E32_NAMES, "E32", "post")
epp_pre  = build_source_rows(EPP_NAMES, "EPP", "pre")
epp_post = build_source_rows(EPP_NAMES, "EPP", "post")

add_sheet(wb, "e32_phase1", E32_HEADER, e32_pre)
add_sheet(wb, "epp_phase1", EPP_HEADER, epp_pre)
add_sheet(wb, "e32_phase2", E32_HEADER, e32_post)
add_sheet(wb, "epp_phase3", EPP_HEADER, epp_post)

# Expected target sheets.
add_sheet(wb, "target_phase0", TGT_HEADER, [])

state = {}
apply_phase(state, "E32", "initial")
apply_phase(state, "EPP", "initial")
add_sheet(wb, "target_phase1", TGT_HEADER, state_to_rows(state))

apply_phase(state, "E32", "delta")
add_sheet(wb, "target_phase2", TGT_HEADER, state_to_rows(state))

apply_phase(state, "EPP", "snapshot")
add_sheet(wb, "target_phase3", TGT_HEADER, state_to_rows(state))


wb.save(OUT)
print(f"wrote {OUT}")
print(f"  e32_phase1 source rows: {len(e32_pre)}")
print(f"  epp_phase1 source rows: {len(epp_pre)}")
print(f"  e32_phase2 source rows: {len(e32_post)}")
print(f"  epp_phase3 source rows: {len(epp_post)}")
